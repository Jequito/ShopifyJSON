import streamlit as st
import requests
import pandas as pd
import html
import re
import io
import time
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Page config ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Shopify Scraper",
    page_icon="🛍️",
    layout="wide",
)

# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean_url(raw: str) -> str:
    raw = raw.strip().rstrip("/")
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    parsed = urlparse(raw)
    return f"{parsed.scheme}://{parsed.netloc}"


def decode_text(text: str) -> str:
    """Decode unicode escapes, HTML entities and strip HTML tags."""
    if not text:
        return ""
    text = re.sub(
        r"\\u([0-9a-fA-F]{4})",
        lambda m: chr(int(m.group(1), 16)),
        text,
    )
    text = html.unescape(text)
    text = text.replace("\u00a0", " ")
    text = re.sub(r"<br\s*/?>" , "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</(p|div|li|h[1-6])>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    })
    return s


def format_duration(seconds: float) -> str:
    if seconds < 1:
        return f"{seconds * 1000:.0f}ms"
    if seconds < 60:
        return f"{seconds:.1f}s"
    minutes, secs = divmod(seconds, 60)
    if minutes < 60:
        return f"{int(minutes)}m {int(secs)}s"
    hours, minutes = divmod(minutes, 60)
    return f"{int(hours)}h {int(minutes)}m {int(secs)}s"


def fetch_json(session: requests.Session, url: str, max_retries: int = 6) -> dict:
    backoff = 2
    for attempt in range(max_retries):
        resp = session.get(url, timeout=20)
        if resp.status_code in (429, 430):
            retry_after = resp.headers.get("Retry-After")
            wait = float(retry_after) if retry_after else backoff
            st.warning(f"Rate limited, waiting {wait:.0f}s before retrying...")
            time.sleep(wait)
            backoff = min(backoff * 2, 30)
            continue
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError(f"Gave up after {max_retries} retries on {url}")

# ─── Products ─────────────────────────────────────────────────────────────────

def scrape_products(base_url: str, limit: int, progress_bar, status_text) -> list[dict]:
    all_rows = []
    page = 1

    with make_session() as session:
        while True:
            status_text.text(f"Products, fetching page {page}... ({len(all_rows)} variants so far)")
            data = fetch_json(session, f"{base_url}/products.json?limit={limit}&page={page}")
            products = data.get("products", [])
            if not products:
                break

            for p in products:
                raw_html = p.get("body_html", "") or ""
                description = decode_text(raw_html)

                base_info = {
                    "Product ID":           p.get("id", ""),
                    "Title":                p.get("title", ""),
                    "Vendor":               p.get("vendor", ""),
                    "Product Type":         p.get("product_type", ""),
                    "Tags":                 ", ".join(p.get("tags", [])),
                    "Description":          description,
                    "Description (HTML)":   raw_html,
                    "Handle":               p.get("handle", ""),
                    "Published At":         p.get("published_at", ""),
                    "Created At":           p.get("created_at", ""),
                    "Updated At":           p.get("updated_at", ""),
                }

                variants = p.get("variants", [])
                images   = p.get("images", [])
                first_image = images[0].get("src", "") if images else ""
                options  = p.get("options", [])

                def opt_name(i):
                    return options[i].get("name", "") if len(options) > i else ""

                if variants:
                    for v in variants:
                        # quantity_rule holds min/max/increment when set
                        qty_rule = v.get("quantity_rule") or {}

                        row = {**base_info, **{
                            "Variant ID":                   v.get("id", ""),
                            "Variant Title":                v.get("title", ""),
                            "SKU":                          v.get("sku", ""),
                            # ── Barcode / GTIN / MPN ──────────────────────────
                            # Shopify stores barcode, EAN, UPC, GTIN and MPN all
                            # in this single field -- whichever the merchant entered.
                            "Barcode / GTIN / MPN":         v.get("barcode") or "",
                            # ── Pricing ───────────────────────────────────────
                            "Price":                        v.get("price", ""),
                            "Price Currency":               v.get("price_currency", ""),
                            "Compare At Price":             v.get("compare_at_price", ""),
                            "Compare At Price Currency":    v.get("compare_at_price_currency", ""),
                            # ── Availability ──────────────────────────────────
                            "Available":                    v.get("available", ""),
                            "Inventory Quantity":           v.get("inventory_quantity", ""),
                            # ── Quantity rules ────────────────────────────────
                            "Min Order Qty":                qty_rule.get("min", ""),
                            "Max Order Qty":                qty_rule.get("max", ""),
                            "Order Qty Increment":          qty_rule.get("increment", ""),
                            # ── Weight / shipping ─────────────────────────────
                            "Weight":                       v.get("weight", ""),
                            "Weight Unit":                  v.get("weight_unit", ""),
                            "Weight (g)":                   v.get("grams", ""),
                            "Requires Shipping":            v.get("requires_shipping", ""),
                            "Taxable":                      v.get("taxable", ""),
                            # ── Options ───────────────────────────────────────
                            "Option1 Name":                 opt_name(0),
                            "Option1 Value":                v.get("option1", ""),
                            "Option2 Name":                 opt_name(1),
                            "Option2 Value":                v.get("option2", ""),
                            "Option3 Name":                 opt_name(2),
                            "Option3 Value":                v.get("option3", ""),
                            # ── Image / URL ───────────────────────────────────
                            "Image URL":    (v.get("featured_image") or {}).get("src", "") or first_image,
                            "Product URL":  f"{base_url}/products/{p.get('handle', '')}",
                        }}
                        all_rows.append(row)
                else:
                    base_info.update({
                        "Variant ID": "", "Variant Title": "", "SKU": "",
                        "Barcode / GTIN / MPN": "",
                        "Price": "", "Price Currency": "",
                        "Compare At Price": "", "Compare At Price Currency": "",
                        "Available": "", "Inventory Quantity": "",
                        "Min Order Qty": "", "Max Order Qty": "", "Order Qty Increment": "",
                        "Weight": "", "Weight Unit": "", "Weight (g)": "",
                        "Requires Shipping": "", "Taxable": "",
                        "Option1 Name": "", "Option1 Value": "",
                        "Option2 Name": "", "Option2 Value": "",
                        "Option3 Name": "", "Option3 Value": "",
                        "Image URL": first_image,
                        "Product URL": f"{base_url}/products/{p.get('handle', '')}",
                    })
                    all_rows.append(base_info)

            progress_bar.progress(min(page / 50, 0.95))
            page += 1
            time.sleep(0.3)

    progress_bar.progress(1.0)
    return all_rows

# ─── Collections ──────────────────────────────────────────────────────────────

def scrape_collections(base_url: str, progress_bar, status_text) -> list[dict]:
    all_rows = []
    page = 1
    LIMIT = 30

    with make_session() as session:
        while True:
            status_text.text(f"Collections, fetching page {page}... ({len(all_rows)} collections so far)")
            data = fetch_json(session, f"{base_url}/collections.json?limit={LIMIT}&page={page}")
            collections = data.get("collections", [])
            if not collections:
                break

            for c in collections:
                handle   = c.get("handle", "")
                raw_html = c.get("description", "") or ""
                description = decode_text(raw_html)
                image_src   = (c.get("image") or {}).get("src", "")

                all_rows.append({
                    "Collection ID":    c.get("id", ""),
                    "Title":            c.get("title", ""),
                    "Handle":           handle,
                    "Description":      description,
                    "Description (HTML)": raw_html,
                    "Published At":     c.get("published_at", ""),
                    "Updated At":       c.get("updated_at", ""),
                    "Sort Order":       c.get("sort_order", ""),
                    "Template Suffix":  c.get("template_suffix", ""),
                    "Product Count":    c.get("products_count", 0),
                    "Image URL":        image_src,
                    "Collection URL":   f"{base_url}/collections/{handle}",
                })

            progress_bar.progress(min(page / 10, 0.95))
            page += 1
            time.sleep(0.4)

    progress_bar.progress(1.0)
    return all_rows

# ─── Excel builder ────────────────────────────────────────────────────────────

def build_xlsx(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font    = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", start_color="EBF1F8")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = cell_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── UI ───────────────────────────────────────────────────────────────────────

st.title("🛍️ Shopify Scraper")
st.caption("Scrape products and collections from any public Shopify store.")

with st.sidebar:
    st.header("⚙️ Settings")
    product_limit = st.slider("Products per page (max 250)", 50, 250, 250, step=50)
    st.caption("Collections are always fetched at 30 per page (Shopify hard limit).")
    st.divider()
    st.markdown(
        "**Endpoints used**\n\n"
        "- `/products.json?limit=250&page=N`\n"
        "- `/collections.json?limit=30&page=N`\n\n"
        "Pages until an empty array is returned."
    )

store_url = st.text_input(
    "Shopify Store URL",
    placeholder="https://example.myshopify.com",
)

mode = st.radio(
    "What would you like to scrape?",
    ["📦 Products", "🗂️ Collections", "📦 + 🗂️ Both"],
    horizontal=True,
)

scrape_btn = st.button("🔍 Scrape", type="primary", use_container_width=True)

for key in ("df_products", "df_collections", "store"):
    if key not in st.session_state:
        st.session_state[key] = None

# ─── Scraping ─────────────────────────────────────────────────────────────────

if scrape_btn and store_url:
    cleaned = clean_url(store_url)
    st.session_state.store = cleaned

    with st.spinner("Connecting..."):
        try:
            r = requests.get(f"{cleaned}/products.json?limit=1&page=1", timeout=10)
            r.raise_for_status()
        except Exception as e:
            st.error(f"❌ Could not reach `{cleaned}`, is this a public Shopify store?\n\n`{e}`")
            st.stop()

    st.info(f"✅ Connected to **{cleaned}**")

    do_products   = mode in ["📦 Products",    "📦 + 🗂️ Both"]
    do_collections = mode in ["🗂️ Collections", "📦 + 🗂️ Both"]

    overall_start = time.perf_counter()

    if do_products:
        st.subheader("Scraping products...")
        prog_p = st.progress(0)
        stat_p = st.empty()
        try:
            t0   = time.perf_counter()
            rows = scrape_products(cleaned, product_limit, prog_p, stat_p)
            elapsed = time.perf_counter() - t0
            st.session_state.df_products = pd.DataFrame(rows)
            stat_p.text(
                f"✅ {len(rows)} variants across "
                f"{st.session_state.df_products['Product ID'].nunique()} products "
                f"in {format_duration(elapsed)}"
            )
        except Exception as e:
            st.error(f"Products scrape failed: {e}")

    if do_collections:
        st.subheader("Scraping collections...")
        prog_c = st.progress(0)
        stat_c = st.empty()
        try:
            t0   = time.perf_counter()
            rows = scrape_collections(cleaned, prog_c, stat_c)
            elapsed = time.perf_counter() - t0
            st.session_state.df_collections = pd.DataFrame(rows)
            stat_c.text(f"✅ {len(rows)} collections found in {format_duration(elapsed)}")
        except Exception as e:
            st.error(f"Collections scrape failed: {e}")

    total_elapsed = time.perf_counter() - overall_start
    st.success(f"🎉 Scrape complete in **{format_duration(total_elapsed)}**")

elif scrape_btn and not store_url:
    st.warning("Please enter a store URL first.")

# ─── Results ──────────────────────────────────────────────────────────────────

has_products    = st.session_state.df_products    is not None and not st.session_state.df_products.empty
has_collections = st.session_state.df_collections is not None and not st.session_state.df_collections.empty

if has_products or has_collections:
    st.divider()
    store_slug = urlparse(st.session_state.store or "store").netloc.replace(".", "_")

    tab_labels = []
    if has_products:    tab_labels.append("📦 Products")
    if has_collections: tab_labels.append("🗂️ Collections")

    tabs    = st.tabs(tab_labels)
    tab_map = dict(zip(tab_labels, tabs))

    # ── Products tab ──────────────────────────────────────────────────────────

    if has_products:
        with tab_map["📦 Products"]:
            df = st.session_state.df_products
            st.subheader(f"{df['Product ID'].nunique():,} products · {len(df):,} variants")

            with st.expander("🔎 Filter", expanded=False):
                c1, c2, c3 = st.columns(3)
                search  = c1.text_input("Search title / description", key="p_search")
                vendors = ["All"] + sorted(df["Vendor"].dropna().unique().tolist())
                sel_v   = c2.selectbox("Vendor", vendors, key="p_vendor")
                types   = ["All"] + sorted(df["Product Type"].dropna().unique().tolist())
                sel_t   = c3.selectbox("Product Type", types, key="p_type")
                show_una = st.checkbox("Show unavailable variants", value=True, key="p_una")

            fdf = df.copy()
            if search:
                fdf = fdf[
                    fdf["Title"].str.contains(search, case=False, na=False) |
                    fdf["Description"].str.contains(search, case=False, na=False)
                ]
            if sel_v != "All": fdf = fdf[fdf["Vendor"] == sel_v]
            if sel_t != "All": fdf = fdf[fdf["Product Type"] == sel_t]
            if not show_una:   fdf = fdf[fdf["Available"] != False]

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Products", fdf["Product ID"].nunique())
            m2.metric("Variants", len(fdf))
            m3.metric("Vendors",  fdf["Vendor"].nunique())
            m4.metric("Types",    fdf["Product Type"].nunique())

            display_cols = [c for c in fdf.columns if c not in ("Description", "Description (HTML)")]
            st.dataframe(
                fdf[display_cols],
                use_container_width=True,
                height=500,
                column_config={
                    "Product URL":          st.column_config.LinkColumn("Product URL"),
                    "Image URL":            st.column_config.ImageColumn("Image", width="small"),
                    "Available":            st.column_config.CheckboxColumn("Available"),
                    "Price":                st.column_config.TextColumn("Price"),
                    "Barcode / GTIN / MPN": st.column_config.TextColumn("Barcode / GTIN / MPN"),
                },
            )

            st.divider()
            ca, cb = st.columns(2)
            ca.download_button(
                "⬇️ Excel (.xlsx)",
                data=build_xlsx(fdf, "Products"),
                file_name=f"{store_slug}_products.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            cb.download_button(
                "⬇️ CSV (.csv)",
                data=fdf.to_csv(index=False).encode("utf-8"),
                file_name=f"{store_slug}_products.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── Collections tab ───────────────────────────────────────────────────────

    if has_collections:
        with tab_map["🗂️ Collections"]:
            df = st.session_state.df_collections
            st.subheader(f"{len(df):,} collections found")

            with st.expander("🔎 Filter", expanded=False):
                c1, c2    = st.columns(2)
                search_c  = c1.text_input("Search title / description", key="c_search")
                sort_opts = ["All"] + sorted(df["Sort Order"].dropna().unique().tolist())
                sel_sort  = c2.selectbox("Sort Order", sort_opts, key="c_sort")

            fdf = df.copy()
            if search_c:
                fdf = fdf[
                    fdf["Title"].str.contains(search_c, case=False, na=False) |
                    fdf["Description"].str.contains(search_c, case=False, na=False)
                ]
            if sel_sort != "All":
                fdf = fdf[fdf["Sort Order"] == sel_sort]

            m1, m2, m3 = st.columns(3)
            m1.metric("Collections",    len(fdf))
            m2.metric("Total Products", int(fdf["Product Count"].sum()))
            m3.metric("Avg Products",   f"{fdf['Product Count'].mean():.1f}")

            display_cols = [c for c in fdf.columns if c not in ("Description", "Description (HTML)", "Image URL")]
            st.dataframe(
                fdf[display_cols],
                use_container_width=True,
                height=500,
                column_config={
                    "Collection URL": st.column_config.LinkColumn("Collection URL"),
                    "Product Count":  st.column_config.NumberColumn("# Products"),
                },
            )

            st.divider()
            ca, cb = st.columns(2)
            ca.download_button(
                "⬇️ Excel (.xlsx)",
                data=build_xlsx(fdf, "Collections"),
                file_name=f"{store_slug}_collections.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            cb.download_button(
                "⬇️ CSV (.csv)",
                data=fdf.to_csv(index=False).encode("utf-8"),
                file_name=f"{store_slug}_collections.csv",
                mime="text/csv",
                use_container_width=True,
            )
